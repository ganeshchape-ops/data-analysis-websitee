"""
Unit and Integration Test Suite for Student Data Analysis
Validates mathematical calculations, data cleaning, edge-case resilience,
grading accuracy, rankings, visualizations, and Excel generation.
"""

import os
import unittest
import numpy as np
import pandas as pd
import openpyxl

from utils.data_loader import load_raw_data, detect_column_mapping, apply_column_mapping
from utils.data_cleaning import clean_student_data
from utils.analysis import (
    perform_full_analysis,
    perform_student_level_analysis,
    perform_overall_analysis,
    perform_section_analysis,
    perform_subject_analysis,
    calculate_grade
)
from utils.insights import generate_automated_insights
from utils.excel_report import generate_excel_report
from utils.visualization import generate_all_visualizations, prepare_chartjs_data


class TestStudentDataAnalysis(unittest.TestCase):

    def setUp(self):
        # Create a deterministic mock DataFrame with 6 subjects
        self.mock_data = {
            "Student ID": ["S01", "S02", "S03", "S04", "S05", "S06"],
            "Student Name": ["Alice", "Bob", "Charlie", "Diana", "Evan", "Fiona"],
            "Section": ["A", "A", "A", "B", "B", "B"],
            "Subject 1": [90, 75, 45, 85, 30, 95],
            "Subject 2": [85, 70, 50, 80, 35, 90],
            "Subject 3": [88, 72, 48, 82, 40, 92],
            "Subject 4": [92, 78, 55, 88, 25, 96],
            "Subject 5": [94, 80, 60, 86, 38, 98],
            "Subject 6": [96, 82, 62, 90, 42, 100],
        }
        self.df_raw = pd.DataFrame(self.mock_data)
        
        # Mapping setup
        self.mapping_info = detect_column_mapping(self.df_raw)
        self.std_df, self.display_names = apply_column_mapping(self.df_raw, self.mapping_info["mapping"])
        self.subject_keys = self.display_names["_subject_keys"]

    def test_grade_classification(self):
        """Test grade calculation at boundary points."""
        self.assertEqual(calculate_grade(100.0), "A+")
        self.assertEqual(calculate_grade(90.0), "A+")
        self.assertEqual(calculate_grade(89.9), "A")
        self.assertEqual(calculate_grade(80.0), "A")
        self.assertEqual(calculate_grade(79.9), "B")
        self.assertEqual(calculate_grade(70.0), "B")
        self.assertEqual(calculate_grade(69.9), "C")
        self.assertEqual(calculate_grade(60.0), "C")
        self.assertEqual(calculate_grade(59.9), "D")
        self.assertEqual(calculate_grade(50.0), "D")
        self.assertEqual(calculate_grade(49.9), "E")
        self.assertEqual(calculate_grade(40.0), "E")
        self.assertEqual(calculate_grade(39.9), "F")
        self.assertEqual(calculate_grade(0.0), "F")

    def test_mathematical_totals_and_averages(self):
        """Verify exact mathematical correctness of Total Marks, Average, and Percentage."""
        student_df = perform_student_level_analysis(
            self.std_df, self.subject_keys, self.display_names, pass_mark=40.0, max_subject_mark=100.0
        )

        # Alice: 90+85+88+92+94+96 = 545
        alice_row = student_df[student_df["student_name"] == "Alice"].iloc[0]
        self.assertEqual(alice_row["total_marks"], 545.0)
        self.assertAlmostEqual(alice_row["average_marks"], 545.0 / 6.0, places=2)
        self.assertAlmostEqual(alice_row["percentage"], (545.0 / 6.0), places=2)
        self.assertEqual(alice_row["grade"], "A+")
        self.assertEqual(alice_row["pass_status"], "Pass")

        # Evan: 30, 35, 40, 25, 38, 42 -> 4 subjects failed (< 40) -> Total = 210, Avg = 35.0
        evan_row = student_df[student_df["student_name"] == "Evan"].iloc[0]
        self.assertEqual(evan_row["total_marks"], 210.0)
        self.assertEqual(evan_row["average_marks"], 35.0)
        self.assertEqual(evan_row["percentage"], 35.0)
        self.assertEqual(evan_row["grade"], "F")
        self.assertEqual(evan_row["pass_status"], "Fail")
        self.assertEqual(evan_row["failed_subjects_count"], 4)

    def test_missing_marks_handling_no_zero_assumption(self):
        """Ensure missing marks are NOT treated as 0 and denominator is strictly valid subjects."""
        df_with_nan = self.std_df.copy()
        # Set Subject 6 of Bob to NaN
        df_with_nan.loc[df_with_nan["student_name"] == "Bob", "subject_6"] = np.nan

        student_df = perform_student_level_analysis(
            df_with_nan, self.subject_keys, self.display_names, pass_mark=40.0
        )
        bob_row = student_df[student_df["student_name"] == "Bob"].iloc[0]

        # Bob valid marks: 75, 70, 72, 78, 80 -> sum = 375 across 5 valid subjects
        self.assertEqual(bob_row["valid_subjects"], 5)
        self.assertEqual(bob_row["total_marks"], 375.0)
        self.assertEqual(bob_row["average_marks"], 75.0)
        self.assertEqual(bob_row["percentage"], 75.0)
        self.assertEqual(bob_row["grade"], "B")

    def test_section_reconciliation(self):
        """Verify section-level metrics reconcile mathematically with student roster records."""
        results = perform_full_analysis(self.std_df, self.subject_keys, self.display_names)
        sec_df = results["section_summary_df"]
        student_df = results["student_df"]

        for _, sec_row in sec_df.iterrows():
            sec_name = sec_row["section"]
            sec_students = student_df[student_df["section"] == sec_name]
            
            # Check student count
            self.assertEqual(sec_row["students_count"], len(sec_students))
            
            # Check sum of totals
            expected_sum = sec_students["total_marks"].sum()
            self.assertAlmostEqual(sec_row["total_marks_sum"], expected_sum, places=2)
            
            # Check pass count
            expected_pass = (sec_students["pass_status"] == "Pass").sum()
            self.assertEqual(sec_row["pass_count"], expected_pass)

    def test_subject_analysis_reconciliation(self):
        """Verify individual 6-subject analysis metrics reconcile with column values."""
        results = perform_full_analysis(self.std_df, self.subject_keys, self.display_names)
        sub_df = results["subject_summary_df"]
        student_df = results["student_df"]

        for _, sub_row in sub_df.iterrows():
            sub_key = sub_row["subject_key"]
            series = student_df[sub_key].dropna()

            self.assertEqual(sub_row["valid_students"], len(series))
            self.assertAlmostEqual(sub_row["average_marks"], series.mean(), places=2)
            self.assertEqual(sub_row["maximum_marks"], series.max())
            self.assertEqual(sub_row["minimum_marks"], series.min())

    def test_data_cleaning_with_anomalies(self):
        """Verify cleaning correctly flags non-numeric strings, out-of-range marks, and duplicates."""
        dirty_data = {
            "Student ID": ["S01", "S01", "S02", "S03"],  # S01 duplicate
            "Student Name": ["Alice", "Alice", "Bob", "Charlie"],
            "Section": [" sec a ", "SEC A", "Section B ", "Section C"],
            "Sub1": [80, 80, "Absent", 95],  # Non-numeric string
            "Sub2": [85, 85, 75, 120],       # Out of range 120
            "Sub3": [90, 90, 80, -10],       # Out of range -10
            "Sub4": [70, 70, 65, 80],
            "Sub5": [75, 75, 70, 85],
            "Sub6": [80, 80, 75, 90]
        }
        df_dirty = pd.DataFrame(dirty_data)
        mapping_info = detect_column_mapping(df_dirty)
        std_df, display_names = apply_column_mapping(df_dirty, mapping_info["mapping"])
        
        cleaned_df, quality = clean_student_data(
            std_df, display_names["_subject_keys"], display_names, min_mark=0.0, max_mark=100.0, remove_duplicates=True
        )

        # Should remove 1 duplicate S01
        self.assertEqual(len(cleaned_df), 3)
        self.assertEqual(quality["duplicate_ids"], 1)
        self.assertEqual(quality["invalid_marks_count"], 1)  # 'Absent'
        self.assertEqual(quality["out_of_range_count"], 2)   # 120 and -10
        # Check that sections are standardized
        self.assertEqual(list(cleaned_df["section"].unique()), ["SEC A", "SECTION B", "SECTION C"])

    def test_excel_report_generation(self):
        """Verify that openpyxl generates an 11-sheet workbook without errors."""
        results = perform_full_analysis(self.std_df, self.subject_keys, self.display_names)
        cleaned_df, quality = clean_student_data(self.std_df, self.subject_keys, self.display_names)
        insights = generate_automated_insights(results, quality)
        
        output_file = "tests/test_output_report.xlsx"
        os.makedirs("tests", exist_ok=True)
        
        report_path = generate_excel_report(
            raw_df=self.df_raw,
            analysis_results=results,
            quality_summary=quality,
            insights=insights,
            output_path=output_file
        )

        self.assertTrue(os.path.exists(report_path))

        # Inspect generated workbook sheets
        wb = openpyxl.load_workbook(report_path)
        expected_sheets = [
            "Dashboard", "Raw_Data", "Cleaned_Data", "Data_Quality", "Overall_Summary",
            "Section_Analysis", "Subject_Analysis", "Section_Subject",
            "Student_Performance", "Top_Students", "Insights"
        ]
        self.assertEqual(wb.sheetnames, expected_sheets)
        
        # Verify dashboard sheet has content
        ws_dash = wb["Dashboard"]
        self.assertIsNotNone(ws_dash.cell(row=1, column=1).value)
        wb.close()
        
        # Clean up
        if os.path.exists(output_file):
            os.remove(output_file)

    def test_visualizations_generation(self):
        """Verify that all 8 charts generate valid base64 image strings."""
        results = perform_full_analysis(self.std_df, self.subject_keys, self.display_names)
        charts_b64 = generate_all_visualizations(results)
        
        expected_chart_keys = [
            "section_avg", "subject_avg", "section_pass", "subject_pass",
            "section_subject_heatmap", "score_dist", "grade_dist", "top_students"
        ]
        for k in expected_chart_keys:
            self.assertIn(k, charts_b64)
            self.assertTrue(len(charts_b64[k]) > 50)  # Valid non-empty base64 string

        chartjs_data = prepare_chartjs_data(results)
        self.assertIn("section_labels", chartjs_data)
        self.assertIn("grade_counts", chartjs_data)


if __name__ == "__main__":
    unittest.main()

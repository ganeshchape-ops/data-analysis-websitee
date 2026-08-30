"""
Excel Report Generator Module
Creates an executive-formatted 11-sheet workbook using OpenPyXL with professional
styling, freeze panes, KPI summary blocks, auto-fit columns, and conditional styling.
"""

import os
from datetime import datetime
from typing import Dict, List, Any, Optional
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Styling Palettes
NAVY_HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
ACCENT_HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
ZEBRA_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
KPI_BG_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

PASS_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
PASS_FONT = Font(name="Segoe UI", size=10, bold=True, color="065F46")

FAIL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
FAIL_FONT = Font(name="Segoe UI", size=10, bold=True, color="991B1B")

WARNING_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

FONT_HEADER = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
FONT_TITLE = Font(name="Segoe UI", size=15, bold=True, color="1E3A8A")
FONT_SUBTITLE = Font(name="Segoe UI", size=10, italic=True, color="64748B")
FONT_REGULAR = Font(name="Segoe UI", size=10, color="1E293B")
FONT_BOLD = Font(name="Segoe UI", size=10, bold=True, color="1E293B")
FONT_KPI_VALUE = Font(name="Segoe UI", size=18, bold=True, color="1E3A8A")
FONT_KPI_LABEL = Font(name="Segoe UI", size=9, bold=True, color="64748B")

BORDER_THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1")
)

BORDER_BOTTOM_DOUBLE = Border(
    bottom=Side(style="double", color="1E3A8A"),
    top=Side(style="thin", color="CBD5E1")
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def _apply_sheet_settings(ws):
    """Enables gridlines and standard row heights."""
    ws.views.sheetView[0].showGridLines = True
    ws.sheet_properties.tabColor = "1E3A8A"


def _autofit_columns(ws, min_width=12, max_width=45):
    """Calculates optimal column width with padding."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val = str(cell.value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, min_width), max_width)


def _write_table_header(ws, start_row: int, headers: List[str], fill=NAVY_HEADER_FILL):
    """Writes a styled table header row."""
    ws.row_dimensions[start_row].height = 26
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = fill
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN


def _add_title_block(ws, title: str, subtitle: str = ""):
    """Adds standard report title block."""
    ws.row_dimensions[1].height = 30
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = FONT_TITLE
    title_cell.alignment = ALIGN_LEFT
    
    if subtitle:
        ws.row_dimensions[2].height = 18
        sub_cell = ws.cell(row=2, column=1, value=subtitle)
        sub_cell.font = FONT_SUBTITLE
        sub_cell.alignment = ALIGN_LEFT


def _build_dashboard_sheet(ws, overall: Dict[str, Any], sec_df: pd.DataFrame, sub_df: pd.DataFrame, quality: Dict[str, Any]):
    """Sheet 1: Executive KPI Dashboard"""
    _apply_sheet_settings(ws)
    _add_title_block(ws, "Student Performance Executive Dashboard", f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Confidential Analytics")

    # KPI Block 1 (Row 4-5)
    kpis = [
        ("TOTAL STUDENTS", f"{overall.get('total_students', 0):,}", "Active Cohort Size", 1),
        ("TOTAL SECTIONS", f"{len(sec_df)}", "Assigned Cohorts", 3),
        ("SUBJECTS ANALYZED", f"{len(sub_df)}", "Evaluated Subjects", 5),
        ("OVERALL AVERAGE", f"{overall.get('overall_percentage', 0.0):.1f}%", "Cohort Mean Score", 7),
        ("PASS PERCENTAGE", f"{overall.get('pass_percentage', 0.0):.1f}%", "Passed All Subjects", 9),
    ]

    for label, val, note, col in kpis:
        # Header / Label
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
        c_lbl = ws.cell(row=4, column=col, value=label)
        c_lbl.font = FONT_KPI_LABEL
        c_lbl.alignment = ALIGN_CENTER
        c_lbl.fill = KPI_BG_FILL

        # Value
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)
        c_val = ws.cell(row=5, column=col, value=val)
        c_val.font = FONT_KPI_VALUE
        c_val.alignment = ALIGN_CENTER
        c_val.fill = KPI_BG_FILL

        # Border
        for r in range(4, 6):
            for c in range(col, col+2):
                ws.cell(row=r, column=c).border = BORDER_THIN

    # Section Summary Snapshot (Row 8)
    ws.cell(row=7, column=1, value="Section Performance Summary").font = FONT_BOLD
    sec_headers = ["Section", "Students", "Average %", "Pass Rate %", "Highest %", "Lowest %", "Consistency (Std Dev)"]
    _write_table_header(ws, 8, sec_headers, ACCENT_HEADER_FILL)

    row_num = 9
    for _, r in sec_df.iterrows():
        ws.row_dimensions[row_num].height = 20
        ws.cell(row=row_num, column=1, value=str(r["section"])).alignment = ALIGN_CENTER
        ws.cell(row=row_num, column=2, value=int(r["students_count"])).alignment = ALIGN_CENTER
        
        c_avg = ws.cell(row=row_num, column=3, value=r["average_percentage"] / 100.0)
        c_avg.number_format = "0.0%"
        c_avg.alignment = ALIGN_RIGHT
        
        c_pass = ws.cell(row=row_num, column=4, value=r["pass_percentage"] / 100.0)
        c_pass.number_format = "0.0%"
        c_pass.alignment = ALIGN_RIGHT
        if r["pass_percentage"] >= 80:
            c_pass.fill = PASS_FILL
            c_pass.font = PASS_FONT
        elif r["pass_percentage"] < 60:
            c_pass.fill = FAIL_FILL
            c_pass.font = FAIL_FONT
            
        c_hi = ws.cell(row=row_num, column=5, value=r["highest_percentage"] / 100.0)
        c_hi.number_format = "0.0%"
        c_hi.alignment = ALIGN_RIGHT
        
        c_lo = ws.cell(row=row_num, column=6, value=r["lowest_percentage"] / 100.0)
        c_lo.number_format = "0.0%"
        c_lo.alignment = ALIGN_RIGHT

        c_sd = ws.cell(row=row_num, column=7, value=r["std_dev"])
        c_sd.number_format = "0.00"
        c_sd.alignment = ALIGN_RIGHT

        for c in range(1, 8):
            ws.cell(row=row_num, column=c).border = BORDER_THIN
            if (row_num % 2) == 0 and not ws.cell(row=row_num, column=c).fill.start_color.rgb:
                ws.cell(row=row_num, column=c).fill = ZEBRA_FILL
        row_num += 1

    # Subject Performance Snapshot
    sub_start_row = row_num + 2
    ws.cell(row=sub_start_row - 1, column=1, value="Subject Performance Summary (6 Subjects)").font = FONT_BOLD
    sub_headers = ["Subject", "Valid Students", "Average Marks", "Median", "Pass Rate %", "Max Marks", "Min Marks"]
    _write_table_header(ws, sub_start_row, sub_headers, ACCENT_HEADER_FILL)

    row_num = sub_start_row + 1
    for _, r in sub_df.iterrows():
        ws.row_dimensions[row_num].height = 20
        ws.cell(row=row_num, column=1, value=str(r["subject_name"])).alignment = ALIGN_LEFT
        ws.cell(row=row_num, column=2, value=int(r["valid_students"])).alignment = ALIGN_CENTER
        
        c_avg = ws.cell(row=row_num, column=3, value=r["average_marks"])
        c_avg.number_format = "0.00"
        c_avg.alignment = ALIGN_RIGHT
        
        c_med = ws.cell(row=row_num, column=4, value=r["median_marks"])
        c_med.number_format = "0.00"
        c_med.alignment = ALIGN_RIGHT
        
        c_pass = ws.cell(row=row_num, column=5, value=r["pass_percentage"] / 100.0)
        c_pass.number_format = "0.0%"
        c_pass.alignment = ALIGN_RIGHT
        if r["pass_percentage"] >= 85:
            c_pass.fill = PASS_FILL
            c_pass.font = PASS_FONT
        elif r["pass_percentage"] < 70:
            c_pass.fill = FAIL_FILL
            c_pass.font = FAIL_FONT

        c_max = ws.cell(row=row_num, column=6, value=r["maximum_marks"])
        c_max.number_format = "0.0"
        c_max.alignment = ALIGN_RIGHT

        c_min = ws.cell(row=row_num, column=7, value=r["minimum_marks"])
        c_min.number_format = "0.0"
        c_min.alignment = ALIGN_RIGHT

        for c in range(1, 8):
            ws.cell(row=row_num, column=c).border = BORDER_THIN
            if (row_num % 2) == 0 and not ws.cell(row=row_num, column=c).fill.start_color.rgb:
                ws.cell(row=row_num, column=c).fill = ZEBRA_FILL
        row_num += 1

    _autofit_columns(ws)


def _build_raw_data_sheet(ws, raw_df: pd.DataFrame):
    """Sheet 2: Untouched Raw Uploaded Data"""
    _apply_sheet_settings(ws)
    _add_title_block(ws, "Raw Uploaded Student Records", "Original source data untouched")
    
    headers = list(raw_df.columns)
    _write_table_header(ws, 4, headers)
    ws.freeze_panes = "A5"

    row_num = 5
    for _, row in raw_df.iterrows():
        ws.row_dimensions[row_num].height = 19
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val if not pd.isna(val) else "")
            cell.border = BORDER_THIN
            cell.font = FONT_REGULAR
            if (row_num % 2) == 0:
                cell.fill = ZEBRA_FILL
        row_num += 1

    _autofit_columns(ws)


def _build_cleaned_data_sheet(ws, student_df: pd.DataFrame, subject_keys: List[str], display_names: Dict[str, str]):
    """Sheet 3: Cleaned Dataset with Calculated Attributes"""
    _apply_sheet_settings(ws)
    _add_title_block(ws, "Standardized & Cleaned Student Dataset", "Cleaned records with computed metrics")

    base_cols = ["student_id", "student_name", "section"]
    if "gender" in student_df.columns:
        base_cols.append("gender")
    if "attendance" in student_df.columns:
        base_cols.append("attendance")

    headers = [col.replace("_", " ").title() for col in base_cols]
    for sk in subject_keys:
        headers.append(display_names.get(sk, sk))
    headers.extend(["Valid Subs", "Total Marks", "Average Marks", "Percentage", "Grade", "Rank", "Status"])

    _write_table_header(ws, 4, headers)
    ws.freeze_panes = "A5"

    row_num = 5
    for _, row in student_df.iterrows():
        ws.row_dimensions[row_num].height = 19
        col_idx = 1
        for col_name in base_cols:
            c = ws.cell(row=row_num, column=col_idx, value=row.get(col_name, ""))
            c.alignment = ALIGN_CENTER if col_name in ["student_id", "section", "gender"] else ALIGN_LEFT
            c.border = BORDER_THIN
            c.font = FONT_REGULAR
            col_idx += 1

        for sk in subject_keys:
            val = row.get(sk)
            c = ws.cell(row=row_num, column=col_idx, value=val if not pd.isna(val) else "")
            c.alignment = ALIGN_RIGHT
            c.border = BORDER_THIN
            c.font = FONT_REGULAR
            if not pd.isna(val):
                c.number_format = "0.0"
            col_idx += 1

        # Metrics
        ws.cell(row=row_num, column=col_idx, value=int(row["valid_subjects"])).alignment = ALIGN_CENTER
        ws.cell(row=row_num, column=col_idx).border = BORDER_THIN
        col_idx += 1

        c_tot = ws.cell(row=row_num, column=col_idx, value=row["total_marks"])
        c_tot.number_format = "0.0"
        c_tot.alignment = ALIGN_RIGHT
        c_tot.border = BORDER_THIN
        col_idx += 1

        c_avg = ws.cell(row=row_num, column=col_idx, value=row["average_marks"])
        c_avg.number_format = "0.00"
        c_avg.alignment = ALIGN_RIGHT
        c_avg.border = BORDER_THIN
        col_idx += 1

        c_pct = ws.cell(row=row_num, column=col_idx, value=row["percentage"] / 100.0)
        c_pct.number_format = "0.0%"
        c_pct.alignment = ALIGN_RIGHT
        c_pct.border = BORDER_THIN
        col_idx += 1

        c_grd = ws.cell(row=row_num, column=col_idx, value=row["grade"])
        c_grd.alignment = ALIGN_CENTER
        c_grd.border = BORDER_THIN
        c_grd.font = FONT_BOLD
        col_idx += 1

        c_rnk = ws.cell(row=row_num, column=col_idx, value=int(row["rank"]))
        c_rnk.alignment = ALIGN_CENTER
        c_rnk.border = BORDER_THIN
        c_rnk.font = FONT_BOLD
        col_idx += 1

        c_sts = ws.cell(row=row_num, column=col_idx, value=row["pass_status"])
        c_sts.alignment = ALIGN_CENTER
        c_sts.border = BORDER_THIN
        if row["pass_status"] == "Pass":
            c_sts.fill = PASS_FILL
            c_sts.font = PASS_FONT
        else:
            c_sts.fill = FAIL_FILL
            c_sts.font = FAIL_FONT

        row_num += 1

    _autofit_columns(ws)


def _build_data_quality_sheet(ws, quality: Dict[str, Any]):
    """Sheet 4: Data Quality Audit Report"""
    _apply_sheet_settings(ws)
    _add_title_block(ws, "Data Quality & Health Audit", "Comprehensive inspection log of input records")

    # Metrics Block
    audit_metrics = [
        ("Total Records Ingested", quality.get("total_records_initial", 0)),
        ("Valid Cleaned Records", quality.get("valid_records", 0)),
        ("Duplicate Records Purged", quality.get("duplicate_rows", 0) + quality.get("duplicate_ids", 0)),
        ("Missing Subject Marks", quality.get("missing_marks_count", 0)),
        ("Invalid/Non-Numeric Marks", quality.get("invalid_marks_count", 0)),
        ("Out-of-Range Marks", quality.get("out_of_range_count", 0)),
        ("Overall Data Health Score", f"{quality.get('quality_score', 100):.1f} / 100")
    ]

    _write_table_header(ws, 4, ["Audit Metric", "Value"], ACCENT_HEADER_FILL)
    for idx, (metric, val) in enumerate(audit_metrics, start=5):
        ws.row_dimensions[idx].height = 20
        ws.cell(row=idx, column=1, value=metric).font = FONT_BOLD
        ws.cell(row=idx, column=1).border = BORDER_THIN
        
        c_val = ws.cell(row=idx, column=2, value=val)
        c_val.font = FONT_REGULAR
        c_val.alignment = ALIGN_CENTER
        c_val.border = BORDER_THIN
        if "Health Score" in metric:
            c_val.font = Font(name="Segoe UI", size=11, bold=True, color="1E3A8A")

    # Anomalies Log Table
    anomalies = quality.get("anomalous_cells", [])
    start_anom = len(audit_metrics) + 7
    ws.cell(row=start_anom - 1, column=1, value="Detected Data Issues & Corrections").font = FONT_BOLD
    _write_table_header(ws, start_anom, ["Student ID", "Student Name", "Subject", "Raw Value", "Validation Action / Reason"], NAVY_HEADER_FILL)

    row_num = start_anom + 1
    if anomalies:
        for item in anomalies:
            ws.row_dimensions[row_num].height = 19
            ws.cell(row=row_num, column=1, value=str(item.get("student_id", ""))).alignment = ALIGN_CENTER
            ws.cell(row=row_num, column=2, value=str(item.get("student_name", ""))).alignment = ALIGN_LEFT
            ws.cell(row=row_num, column=3, value=str(item.get("subject", ""))).alignment = ALIGN_CENTER
            ws.cell(row=row_num, column=4, value=str(item.get("raw_value", ""))).alignment = ALIGN_CENTER
            ws.cell(row=row_num, column=5, value=str(item.get("reason", ""))).alignment = ALIGN_LEFT
            
            for c in range(1, 6):
                ws.cell(row=row_num, column=c).border = BORDER_THIN
                ws.cell(row=row_num, column=c).font = FONT_REGULAR
            row_num += 1
    else:
        ws.row_dimensions[row_num].height = 20
        ws.cell(row=row_num, column=1, value="No anomalous or corrupt data points detected. Dataset is 100% clean.").font = FONT_REGULAR
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)

    _autofit_columns(ws)


def _build_overall_summary_sheet(ws, overall: Dict[str, Any]):
    """Sheet 5: Overall Statistical Summary"""
    _apply_sheet_settings(ws)
    _add_title_block(ws, "Overall Cohort Statistical Summary", "Detailed descriptive statistical parameters")

    stats_list = [
        ("Total Students Evaluated", overall.get("total_students", 0), "Number of candidate records"),
        ("Mean Average Score (%)", f"{overall.get('overall_percentage', 0.0):.2f}%", "Arithmetic average percentage"),
        ("Median Percentage (%)", f"{overall.get('median_percentage', 0.0):.2f}%", "50th percentile midpoint"),
        ("Mode Percentage (%)", f"{overall.get('mode_percentage', 'N/A')}%", "Most frequent percentage"),
        ("Highest Percentage Score (%)", f"{overall.get('highest_percentage', 0.0):.2f}%", "Cohort maximum mark"),
        ("Lowest Percentage Score (%)", f"{overall.get('lowest_percentage', 0.0):.2f}%", "Cohort minimum mark"),
        ("Score Range (%)", f"{overall.get('range', 0.0):.2f}%", "Max minus Min spread"),
        ("Interquartile Range (IQR)", f"{overall.get('iqr', 0.0):.2f}%", "Q3 minus Q1 spread"),
        ("Standard Deviation", f"{overall.get('std_dev', 0.0):.2f}", "Dispersion around cohort mean"),
        ("Sample Variance", f"{overall.get('variance', 0.0):.2f}", "Square of standard deviation"),
        ("Total Pass Count", overall.get("pass_count", 0), "Students meeting passing criteria"),
        ("Total Fail Count", overall.get("fail_count", 0), "Students requiring academic support"),
        ("Cohort Pass Rate (%)", f"{overall.get('pass_percentage', 0.0):.2f}%", "Overall pass percentage"),
        ("Cohort Fail Rate (%)", f"{overall.get('fail_percentage', 0.0):.2f}%", "Overall fail percentage")
    ]

    _write_table_header(ws, 4, ["Statistical Parameter", "Value", "Interpretation / Notes"])
    
    for idx, (param, val, desc) in enumerate(stats_list, start=5):
        ws.row_dimensions[idx].height = 21
        ws.cell(row=idx, column=1, value=param).font = FONT_BOLD
        ws.cell(row=idx, column=1).border = BORDER_THIN
        
        c_val = ws.cell(row=idx, column=2, value=val)
        c_val.font = FONT_REGULAR
        c_val.alignment = ALIGN_CENTER
        c_val.border = BORDER_THIN

        c_desc = ws.cell(row=idx, column=3, value=desc)
        c_desc.font = FONT_SUBTITLE
        c_desc.alignment = ALIGN_LEFT
        c_desc.border = BORDER_THIN

        if (idx % 2) == 0:
            ws.cell(row=idx, column=1).fill = ZEBRA_FILL
            ws.cell(row=idx, column=2).fill = ZEBRA_FILL
            ws.cell(row=idx, column=3).fill = ZEBRA_FILL

    _autofit_columns(ws)


def _build_section_analysis_sheet(ws, sec_df: pd.DataFrame, highlights: Dict[str, Any]):
    """Sheet 6: Section-wise Comparative Analysis"""
    _apply_sheet_settings(ws)
    _add_title_block(ws, "Section-Wise Performance Analysis", "Comparative breakdown across classroom sections")

    headers = [
        "Section", "Students", "Total Marks Sum", "Average Marks", "Median Marks",
        "Average %", "Highest Marks", "Lowest Marks", "Pass Count", "Fail Count",
        "Pass Rate %", "Consistency (Std Dev)"
    ]
    _write_table_header(ws, 4, headers)
    ws.freeze_panes = "A5"

    row_num = 5
    for _, r in sec_df.iterrows():
        ws.row_dimensions[row_num].height = 20
        ws.cell(row=row_num, column=1, value=str(r["section"])).alignment = ALIGN_CENTER
        ws.cell(row=row_num, column=2, value=int(r["students_count"])).alignment = ALIGN_CENTER
        
        c_sum = ws.cell(row=row_num, column=3, value=r["total_marks_sum"])
        c_sum.number_format = "#,##0.0"
        c_sum.alignment = ALIGN_RIGHT

        c_avg_m = ws.cell(row=row_num, column=4, value=r["average_marks"])
        c_avg_m.number_format = "0.00"
        c_avg_m.alignment = ALIGN_RIGHT

        c_med = ws.cell(row=row_num, column=5, value=r["median_marks"])
        c_med.number_format = "0.00"
        c_med.alignment = ALIGN_RIGHT

        c_avg_p = ws.cell(row=row_num, column=6, value=r["average_percentage"] / 100.0)
        c_avg_p.number_format = "0.0%"
        c_avg_p.alignment = ALIGN_RIGHT

        c_hi = ws.cell(row=row_num, column=7, value=r["highest_marks"])
        c_hi.number_format = "0.0"
        c_hi.alignment = ALIGN_RIGHT

        c_lo = ws.cell(row=row_num, column=8, value=r["lowest_marks"])
        c_lo.number_format = "0.0"
        c_lo.alignment = ALIGN_RIGHT

        ws.cell(row=row_num, column=9, value=int(r["pass_count"])).alignment = ALIGN_CENTER
        ws.cell(row=row_num, column=10, value=int(r["fail_count"])).alignment = ALIGN_CENTER

        c_pass = ws.cell(row=row_num, column=11, value=r["pass_percentage"] / 100.0)
        c_pass.number_format = "0.0%"
        c_pass.alignment = ALIGN_RIGHT
        if r["pass_percentage"] >= 80.0:
            c_pass.fill = PASS_FILL
            c_pass.font = PASS_FONT
        elif r["pass_percentage"] < 60.0:
            c_pass.fill = FAIL_FILL
            c_pass.font = FAIL_FONT

        c_sd = ws.cell(row=row_num, column=12, value=r["std_dev"])
        c_sd.number_format = "0.00"
        c_sd.alignment = ALIGN_RIGHT

        for c in range(1, 13):
            ws.cell(row=row_num, column=c).border = BORDER_THIN
            if (row_num % 2) == 0 and not ws.cell(row=row_num, column=c).fill.start_color.rgb:
                ws.cell(row=row_num, column=c).fill = ZEBRA_FILL
        row_num += 1

    _autofit_columns(ws)


def _build_subject_analysis_sheet(ws, sub_df: pd.DataFrame, highlights: Dict[str, Any]):
    """Sheet 7: 6 Subject Individual Analysis"""
    _apply_sheet_settings(ws)
    _add_title_block(ws, "Subject-Wise Detailed Analysis (6 Subjects)", "Comprehensive evaluation of individual subject performance")

    headers = [
        "Subject Name", "Valid Students", "Average Marks", "Median Marks",
        "Highest Marks", "Lowest Marks", "Std Dev", "Pass Count", "Fail Count",
        "Pass Rate %", "Performance %"
    ]
    _write_table_header(ws, 4, headers)
    ws.freeze_panes = "A5"

    row_num = 5
    for _, r in sub_df.iterrows():
        ws.row_dimensions[row_num].height = 20
        ws.cell(row=row_num, column=1, value=str(r["subject_name"])).alignment = ALIGN_LEFT
        ws.cell(row=row_num, column=2, value=int(r["valid_students"])).alignment = ALIGN_CENTER

        c_avg = ws.cell(row=row_num, column=3, value=r["average_marks"])
        c_avg.number_format = "0.00"
        c_avg.alignment = ALIGN_RIGHT

        c_med = ws.cell(row=row_num, column=4, value=r["median_marks"])
        c_med.number_format = "0.00"
        c_med.alignment = ALIGN_RIGHT

        c_hi = ws.cell(row=row_num, column=5, value=r["maximum_marks"])
        c_hi.number_format = "0.0"
        c_hi.alignment = ALIGN_RIGHT

        c_lo = ws.cell(row=row_num, column=6, value=r["minimum_marks"])
        c_lo.number_format = "0.0"
        c_lo.alignment = ALIGN_RIGHT

        c_sd = ws.cell(row=row_num, column=7, value=r["std_dev"])
        c_sd.number_format = "0.00"
        c_sd.alignment = ALIGN_RIGHT

        ws.cell(row=row_num, column=8, value=int(r["pass_count"])).alignment = ALIGN_CENTER
        ws.cell(row=row_num, column=9, value=int(r["fail_count"])).alignment = ALIGN_CENTER

        c_pass = ws.cell(row=row_num, column=10, value=r["pass_percentage"] / 100.0)
        c_pass.number_format = "0.0%"
        c_pass.alignment = ALIGN_RIGHT
        if r["pass_percentage"] >= 85.0:
            c_pass.fill = PASS_FILL
            c_pass.font = PASS_FONT
        elif r["pass_percentage"] < 70.0:
            c_pass.fill = FAIL_FILL
            c_pass.font = FAIL_FONT

        c_perf = ws.cell(row=row_num, column=11, value=r["performance_percentage"] / 100.0)
        c_perf.number_format = "0.0%"
        c_perf.alignment = ALIGN_RIGHT

        for c in range(1, 12):
            ws.cell(row=row_num, column=c).border = BORDER_THIN
            if (row_num % 2) == 0 and not ws.cell(row=row_num, column=c).fill.start_color.rgb:
                ws.cell(row=row_num, column=c).fill = ZEBRA_FILL
        row_num += 1

    _autofit_columns(ws)


def _build_section_subject_sheet(ws, avg_matrix_df: pd.DataFrame, pass_matrix_df: pd.DataFrame):
    """Sheet 8: Section x Subject Matrix"""
    _apply_sheet_settings(ws)
    _add_title_block(ws, "Section × Subject Performance Matrix", "Average marks and pass rate distribution across subjects per section")

    # 1. Average Marks Matrix
    ws.cell(row=4, column=1, value="Average Score Matrix (Marks out of 100)").font = FONT_BOLD
    headers = list(avg_matrix_df.columns)
    _write_table_header(ws, 5, headers, ACCENT_HEADER_FILL)

    row_num = 6
    for _, r in avg_matrix_df.iterrows():
        ws.row_dimensions[row_num].height = 20
        ws.cell(row=row_num, column=1, value=str(r["section"])).alignment = ALIGN_CENTER
        for col_idx, col_name in enumerate(headers[1:], start=2):
            val = r[col_name]
            c = ws.cell(row=row_num, column=col_idx, value=val if val is not None else "")
            c.alignment = ALIGN_RIGHT
            c.border = BORDER_THIN
            if val is not None:
                c.number_format = "0.0"
                if val >= 75.0:
                    c.fill = PASS_FILL
                elif val < 50.0:
                    c.fill = FAIL_FILL
        ws.cell(row=row_num, column=1).border = BORDER_THIN
        row_num += 1

    # 2. Pass Rate Matrix
    start_pass = row_num + 2
    ws.cell(row=start_pass - 1, column=1, value="Subject Pass Rate Matrix (%)").font = FONT_BOLD
    _write_table_header(ws, start_pass, headers, NAVY_HEADER_FILL)

    row_num = start_pass + 1
    for _, r in pass_matrix_df.iterrows():
        ws.row_dimensions[row_num].height = 20
        ws.cell(row=row_num, column=1, value=str(r["section"])).alignment = ALIGN_CENTER
        for col_idx, col_name in enumerate(headers[1:], start=2):
            val = r[col_name]
            c = ws.cell(row=row_num, column=col_idx, value=(val / 100.0) if val is not None else "")
            c.alignment = ALIGN_RIGHT
            c.border = BORDER_THIN
            if val is not None:
                c.number_format = "0.0%"
                if val >= 85.0:
                    c.fill = PASS_FILL
                elif val < 70.0:
                    c.fill = FAIL_FILL
        ws.cell(row=row_num, column=1).border = BORDER_THIN
        row_num += 1

    _autofit_columns(ws)


def _build_student_performance_sheet(ws, student_df: pd.DataFrame, subject_keys: List[str], display_names: Dict[str, str]):
    """Sheet 9: Complete Ranked Student Performance Roster"""
    _apply_sheet_settings(ws)
    _add_title_block(ws, "Student Performance Master Roster", "Full ranked cohort records sorted by performance")

    headers = ["Rank", "Student ID", "Student Name", "Section"]
    for sk in subject_keys:
        headers.append(display_names.get(sk, sk))
    headers.extend(["Total Marks", "Average Marks", "Percentage", "Grade", "Status"])

    _write_table_header(ws, 4, headers)
    ws.freeze_panes = "A5"

    df_sorted = student_df.sort_values(by=["rank", "percentage"], ascending=[True, False])

    row_num = 5
    for _, r in df_sorted.iterrows():
        ws.row_dimensions[row_num].height = 19
        ws.cell(row=row_num, column=1, value=int(r["rank"])).alignment = ALIGN_CENTER
        ws.cell(row=row_num, column=2, value=str(r["student_id"])).alignment = ALIGN_CENTER
        ws.cell(row=row_num, column=3, value=str(r["student_name"])).alignment = ALIGN_LEFT
        ws.cell(row=row_num, column=4, value=str(r["section"])).alignment = ALIGN_CENTER

        col_idx = 5
        for sk in subject_keys:
            val = r[sk]
            c = ws.cell(row=row_num, column=col_idx, value=val if not pd.isna(val) else "")
            c.alignment = ALIGN_RIGHT
            c.border = BORDER_THIN
            if not pd.isna(val):
                c.number_format = "0.0"
                if val < 40.0:
                    c.fill = FAIL_FILL
            col_idx += 1

        c_tot = ws.cell(row=row_num, column=col_idx, value=r["total_marks"])
        c_tot.number_format = "0.0"
        c_tot.alignment = ALIGN_RIGHT
        col_idx += 1

        c_avg = ws.cell(row=row_num, column=col_idx, value=r["average_marks"])
        c_avg.number_format = "0.00"
        c_avg.alignment = ALIGN_RIGHT
        col_idx += 1

        c_pct = ws.cell(row=row_num, column=col_idx, value=r["percentage"] / 100.0)
        c_pct.number_format = "0.0%"
        c_pct.alignment = ALIGN_RIGHT
        col_idx += 1

        c_grd = ws.cell(row=row_num, column=col_idx, value=str(r["grade"]))
        c_grd.alignment = ALIGN_CENTER
        c_grd.font = FONT_BOLD
        col_idx += 1

        c_sts = ws.cell(row=row_num, column=col_idx, value=str(r["pass_status"]))
        c_sts.alignment = ALIGN_CENTER
        if r["pass_status"] == "Pass":
            c_sts.fill = PASS_FILL
            c_sts.font = PASS_FONT
        else:
            c_sts.fill = FAIL_FILL
            c_sts.font = FAIL_FONT

        for c in range(1, col_idx + 1):
            ws.cell(row=row_num, column=c).border = BORDER_THIN
            if (row_num % 2) == 0 and not ws.cell(row=row_num, column=c).fill.start_color.rgb:
                ws.cell(row=row_num, column=c).fill = ZEBRA_FILL
        row_num += 1

    _autofit_columns(ws)


def _build_top_students_sheet(ws, top_10_df: pd.DataFrame):
    """Sheet 10: Top 10 High Achievers Leaderboard"""
    _apply_sheet_settings(ws)
    _add_title_block(ws, "Top 10 High Achievers Leaderboard", "Honors list of top performing students")

    headers = ["Honor Rank", "Student ID", "Student Name", "Section", "Total Marks", "Average Marks", "Percentage", "Grade"]
    _write_table_header(ws, 4, headers, ACCENT_HEADER_FILL)

    row_num = 5
    for _, r in top_10_df.iterrows():
        ws.row_dimensions[row_num].height = 22
        rank_badge = f"#{int(r['rank'])}"
        c_rnk = ws.cell(row=row_num, column=1, value=rank_badge)
        c_rnk.alignment = ALIGN_CENTER
        c_rnk.font = FONT_BOLD
        if r["rank"] == 1:
            c_rnk.fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")  # Gold

        ws.cell(row=row_num, column=2, value=str(r["student_id"])).alignment = ALIGN_CENTER
        ws.cell(row=row_num, column=3, value=str(r["student_name"])).alignment = ALIGN_LEFT
        ws.cell(row=row_num, column=4, value=str(r["section"])).alignment = ALIGN_CENTER

        c_tot = ws.cell(row=row_num, column=5, value=r["total_marks"])
        c_tot.number_format = "0.0"
        c_tot.alignment = ALIGN_RIGHT

        c_avg = ws.cell(row=row_num, column=6, value=r["average_marks"])
        c_avg.number_format = "0.00"
        c_avg.alignment = ALIGN_RIGHT

        c_pct = ws.cell(row=row_num, column=7, value=r["percentage"] / 100.0)
        c_pct.number_format = "0.0%"
        c_pct.alignment = ALIGN_RIGHT
        c_pct.font = FONT_BOLD

        c_grd = ws.cell(row=row_num, column=8, value=str(r["grade"]))
        c_grd.alignment = ALIGN_CENTER
        c_grd.font = FONT_BOLD
        c_grd.fill = PASS_FILL

        for c in range(1, 9):
            ws.cell(row=row_num, column=c).border = BORDER_THIN
        row_num += 1

    _autofit_columns(ws)


def _build_insights_sheet(ws, insights: Dict[str, Any]):
    """Sheet 11: Automated Insights & Executive Recommendations"""
    _apply_sheet_settings(ws)
    _add_title_block(ws, "Automated Insights & Strategic Recommendations", "Data-driven observations and academic intervention guidance")

    categories = [
        ("Executive Summary", insights.get("executive_summary", []), NAVY_HEADER_FILL),
        ("Section Performance Benchmarks", insights.get("section_insights", []), ACCENT_HEADER_FILL),
        ("Subject Performance Trends", insights.get("subject_insights", []), NAVY_HEADER_FILL),
        ("Cross Section-Subject Matrix Observations", insights.get("cross_matrix_insights", []), ACCENT_HEADER_FILL),
        ("Student Academic Alerts", insights.get("student_alerts", []), NAVY_HEADER_FILL),
        ("Strategic Recommendations", insights.get("recommendations", []), ACCENT_HEADER_FILL)
    ]

    row_num = 4
    for title, bullet_list, header_fill in categories:
        ws.row_dimensions[row_num].height = 24
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
        c_head = ws.cell(row=row_num, column=1, value=title)
        c_head.font = FONT_HEADER
        c_head.fill = header_fill
        c_head.alignment = ALIGN_LEFT
        row_num += 1

        for bullet in bullet_list:
            ws.row_dimensions[row_num].height = 20
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
            c_txt = ws.cell(row=row_num, column=1, value=f"•  {bullet}")
            c_txt.font = FONT_REGULAR
            c_txt.alignment = ALIGN_LEFT
            row_num += 1

        row_num += 1  # Blank spacing row

    ws.column_dimensions["A"].width = 95


def generate_excel_report(
    raw_df: pd.DataFrame,
    analysis_results: Dict[str, Any],
    quality_summary: Dict[str, Any],
    insights: Dict[str, Any],
    output_path: str
) -> str:
    """
    Generates a professional 11-sheet Excel workbook.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = openpyxl.Workbook()

    # Sheet 1: Dashboard
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    _build_dashboard_sheet(
        ws_dash,
        analysis_results["overall_stats"],
        analysis_results["section_summary_df"],
        analysis_results["subject_summary_df"],
        quality_summary
    )

    # Sheet 2: Raw_Data
    ws_raw = wb.create_sheet(title="Raw_Data")
    _build_raw_data_sheet(ws_raw, raw_df)

    # Sheet 3: Cleaned_Data
    ws_clean = wb.create_sheet(title="Cleaned_Data")
    _build_cleaned_data_sheet(
        ws_clean,
        analysis_results["student_df"],
        analysis_results["subject_keys"],
        analysis_results["display_names"]
    )

    # Sheet 4: Data_Quality
    ws_dq = wb.create_sheet(title="Data_Quality")
    _build_data_quality_sheet(ws_dq, quality_summary)

    # Sheet 5: Overall_Summary
    ws_overall = wb.create_sheet(title="Overall_Summary")
    _build_overall_summary_sheet(ws_overall, analysis_results["overall_stats"])

    # Sheet 6: Section_Analysis
    ws_sec = wb.create_sheet(title="Section_Analysis")
    _build_section_analysis_sheet(ws_sec, analysis_results["section_summary_df"], analysis_results["section_highlights"])

    # Sheet 7: Subject_Analysis
    ws_sub = wb.create_sheet(title="Subject_Analysis")
    _build_subject_analysis_sheet(ws_sub, analysis_results["subject_summary_df"], analysis_results["subject_highlights"])

    # Sheet 8: Section_Subject
    ws_mat = wb.create_sheet(title="Section_Subject")
    _build_section_subject_sheet(ws_mat, analysis_results["average_matrix_df"], analysis_results["pass_matrix_df"])

    # Sheet 9: Student_Performance
    ws_perf = wb.create_sheet(title="Student_Performance")
    _build_student_performance_sheet(
        ws_perf,
        analysis_results["student_df"],
        analysis_results["subject_keys"],
        analysis_results["display_names"]
    )

    # Sheet 10: Top_Students
    ws_top = wb.create_sheet(title="Top_Students")
    _build_top_students_sheet(ws_top, analysis_results["top_10_df"])

    # Sheet 11: Insights
    ws_ins = wb.create_sheet(title="Insights")
    _build_insights_sheet(ws_ins, insights)

    wb.save(output_path)
    return output_path
